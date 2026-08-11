#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Standalone EQ1–EQ32 Daily Optimization workbook builder.

Runtime inputs ONLY:
  1. Input Data.xlsx
  2. Input Core assumptions.xlsx

No template/output workbook is read or required. The script creates all sheets,
headers, formulas, formatting, summaries, and four battery scenarios from scratch.
"""
from __future__ import annotations

import argparse
import sys
import zipfile
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BATTERY_SIZES = (30.0, 20.0, 15.0, 10.0)
HOURLY_SHEET = "Solar_8760_2025"
MONTHLY_SHEET = "Monthly_Profile"
PARAMETERS_SHEET = "Parameters"
GUIDE_SHEET = "EQ_Guide"
SUMMARY_SHEET = "Summary"
FIRST_DATA_ROW = 2
LAST_DATA_ROW = 8761
HOURS_PER_DAY = 24

HEADERS = [
    "Timestamp", "Date", "Month", "Month_Number", "Day", "Hour",
    "Solar_Availability_kW/(m^2)", "Epv _PV,Generation (KWh)",
    "E,base,t Base load Demand_kWh", "Price,spot,t (€/MWh)",
    "ω_t (€/kWh)", "Price_spot (€/kWh)", "Daily Price Min (€/kWh)",
    "Daily Price Max (€/kWh)", "Daily Low Threshold (€/kWh)",
    "Daily High Threshold (€/kWh)", "Daily PV Max (kWh)", "Decision",
    "E_self,n,t EQ1 (kWh)", "E_charge (kWh)", "E_discharge (kWh)",
    "SOC_begin (kWh)", "SOC_end EQ9/12 (kWh)", "E_grid,t EQ2-EQ4 (kWh)",
    "ε_t Elasticity EQ6 (-)", "U_cons EQ7 (€)", "E_actual EQ22 (kWh)",
    "DeltaE (kWh)", "D_shift (€)", "E_bat EQ11 (kWh)", "C_deg EQ10 (€)",
    "E_sell (kWh)", "E_buy (kWh)", "W_spot (€)", "Utility EQ8/29 (€)",
    "Daily U_max EQ8/EQ29 (€)", "SDR_t = E_inj,prosum,t / E_P2P,ij,t (-)",
    "Price_prosumer,ch-inj,t Daily Charge-Weighted Reference (€/kWh)",
    "Price_P2P,t EQ13 (€/kWh)", "E_inj EQ16 (kWh)", "U_prosumer EQ14 (€)",
    "a1 EQ18", "Ltotal/Cgrid", "gamma EQ18", "tau EQ17 (€/kWh)",
    "Green benefit (€)", "SW_total EQ15 (€)", "U_buyer EQ19 (€)",
    "A EQ21 (€)", "Price_eq EQ20 (€/kWh)", "CS EQ23 (€)", "PS EQ24 (€)",
    "LC EQ26 (€)", "QC EQ27 (€)", "SW_P2P EQ25 (€)", "CS_pros EQ28 (€)",
    "CSt1 EQ30 (€)", "Daily Price Peak (€/kWh)", "GS_grid EQ31 (€)",
    "SW_grid EQ32 (€)"
]

PARAMETER_ROWS = [
    ("Utility curvature", "alpha", 0.25, "€/kWh²", "EQ1–EQ8, EQ19–EQ30", "User-selected"),
    ("Load-shifting coefficient", "zeta", 0.20, "€/kWh²", "EQ8, EQ14, EQ21, EQ24–EQ29", "User-selected"),
    ("Battery degradation coefficient", "psi", 0.03, "€/kWh", "EQ10, EQ14, EQ21, EQ24–EQ29", "User-selected"),
    ("Battery capacity", "C_bat", None, "kWh", "EQ2, EQ4, EQ9, EQ12", None),
    ("Maximum charging power", "P_ch_max", 3.0, "kWh/h", "EQ2, EQ11", "1-hour step"),
    ("Maximum discharging power", "P_dis_max", 3.0, "kWh/h", "EQ4, EQ11, EQ28–EQ31", "1-hour step"),
    ("Charging efficiency", "eta_ch", None, "fraction", "EQ9, EQ12", "Core assumptions"),
    ("Discharging efficiency", "eta_dis", None, "fraction", "EQ9, EQ12", "One-way proxy"),
    ("Initial SOC", "SOC_0", 0.90, "fraction of capacity", "EQ9, EQ12", "90%"),
    ("Minimum SOC", "SOC_min", 0.10, "fraction of capacity", "EQ4, EQ9, EQ12", "Operational bound"),
    ("Maximum SOC", "SOC_max", 1.00, "fraction of capacity", "EQ2, EQ9, EQ12", "Operational bound"),
    ("Local supply-demand ratio", "SDR_t", None, "endogenous ratio", "EQ13 / Column AK", "Calculated hourly as E_inj,prosum,t / E_P2P,ij,t; hourly demand in column I is the proxy denominator."),
    ("Baseline network cost", "tau_base", 0.05, "€/kWh", "EQ17", "Lower stated bound"),
    ("Green preference premium", "theta", 0.022, "€/kWh", "EQ15, EQ19, EQ21, EQ23–EQ27", "Given"),
    ("Green guarantee indicator", "rho", 1.0, "fraction", "EQ15, EQ19, EQ21, EQ23–EQ27", "1 = fully green"),
    ("Prosumer marginal PV cost", "b_i", 0.065, "€/kWh", "EQ24–EQ27", "Range midpoint"),
    ("Prosumer quadratic cost", "a_i", 0.025, "€/kWh²", "EQ24–EQ27", "Range midpoint"),
    ("Network scaling coefficient", "a1_min", 0.10, "fraction", "EQ18", "At minimum load"),
    ("Network scaling coefficient", "a1_max", 0.50, "fraction", "EQ18", "At maximum load"),
    ("Low-price threshold share", "Low_share", 1/3, "fraction", "EQ2 rule", "Lower third daily range"),
    ("High-price threshold share", "High_share", 2/3, "fraction", "EQ4 rule", "Upper third daily range"),
]

GUIDE_ROWS = [
    ("EQ1", "E_self,t", "kWh", "S", "MAX(0,MIN(E_base,PV,(omega-Price_kWh)/alpha))", "Physical caps"),
    ("EQ2–EQ4", "E_grid,t under three price cases", "kWh", "X", "Piecewise charging/standby/discharging", "Daily thresholds"),
    ("EQ6", "Dynamic elasticity", "dimensionless", "Y", "-(1/alpha)*Price_kWh/E_grid", "Zero protected"),
    ("EQ7", "Consumption utility", "€", "Z", "omega*E_actual-alpha/2*E_actual^2", None),
    ("EQ8/EQ29", "Utility objective", "€", "AI/AJ", "U_cons-D_shift+W_spot-C_deg", "AJ is daily total"),
    ("EQ9/EQ12", "SOC dynamics", "kWh", "V/W", "SOC_begin+eta_ch*charge-discharge/eta_dis", "Bounded"),
    ("EQ10", "Battery degradation", "€", "AE", "psi*ABS(E_bat)", None),
    ("EQ11", "Battery energy", "kWh", "AD", "charge-discharge", "Positive=charge"),
    ("EQ13", "P2P price", "€/kWh", "AM", "Deficit formula", "Division protected"),
    ("EQ14", "Prosumer utility", "€", "AO", "U_cons-D_shift+injection revenue-C_deg", None),
    ("EQ15", "Social welfare contribution", "€", "AU", "Utility-costs+green benefit-network cost", None),
    ("EQ16", "Prosumer injection", "kWh", "AN", "E_sell, except during charge", None),
    ("EQ17/EQ18", "Network cost, gamma", "€/kWh", "AP:AS", "Load normalization and quadratic cost", None),
    ("EQ19", "Buyer utility", "€", "AV", "Utility-purchase cost+green premium", None),
    ("EQ20/EQ21", "P2P equilibrium price and A", "€/kWh, €", "AW:AX", "A/(2*E_P2P)", "Blank with zero injection"),
    ("EQ22", "Actual demand", "kWh", "AA", "E_self+E_discharge+E_buy", None),
    ("EQ23", "Consumer surplus", "€", "AY", "Stated CS expression", None),
    ("EQ24", "Producer surplus", "€", "AZ", "Direct analytical expression", None),
    ("EQ25", "P2P social welfare", "€", "BC", "Direct analytical expression", None),
    ("EQ26/EQ27", "Linear/quadratic components", "€", "BA:BB", "Stated decomposition", None),
    ("EQ28", "Prosumer-grid CS", "€", "BD", "Piecewise across price cases", None),
    ("EQ30", "High-price CS branch", "€", "BE", "Not calculated", "BE intentionally blank"),
    ("EQ31", "Grid surplus", "€", "BG", "(Price_peak-Price_spot)*discharge", None),
    ("EQ32", "Grid social welfare", "€", "BH", "CS_pros+GS_grid", None),
]

THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FILL = PatternFill("solid", fgColor="5B9BD5")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
FORMULA_FILL = PatternFill("solid", fgColor="E2F0D9")


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Create EQ1–EQ32 workbooks from two input files; no template required.")
    p.add_argument("--input-data", type=Path, default=Path("Input Data.xlsx"))
    p.add_argument("--core-assumptions", type=Path, default=Path("Input Core assumptions.xlsx"))
    p.add_argument("--output-dir", type=Path, default=Path("outputs_EQ1_EQ32"))
    p.add_argument("--zip-output", type=Path, default=Path("Output_EQ1_EQ32_DailyOptimization_4_Batteries.zip"))
    p.add_argument("--battery-sizes", type=float, nargs="+", default=list(BATTERY_SIZES))
    args, unknown = p.parse_known_args()
    return args


def existing_or_fallback(path: Path, patterns: Iterable[str]) -> Path:
    if path.exists():
        return path
    candidates = []
    for pattern in patterns:
        candidates.extend(sorted(Path.cwd().glob(pattern)))
    candidates = [p for p in candidates if p.is_file()]
    if len(candidates) == 1:
        return candidates[0]
    if candidates:
        return candidates[-1]
    raise FileNotFoundError(f"Required file not found: {path}")


def read_core_assumptions(path: Path) -> tuple[float, float, float]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    values = {}
    for row in ws.iter_rows(min_row=1, values_only=True):
        key = str(row[0]).strip().lower() if row and row[0] is not None else ""
        if key:
            values[key] = row[1] if len(row) > 1 else None
    pv_area = float(values.get("pv area", 50))
    default_battery = float(values.get("battery energy", 30))
    eta = float(values.get("battery efficiency", 0.92))
    return pv_area, default_battery, eta


def copy_input_sheet_values(source_ws, target_ws, max_col: int | None = None) -> None:
    max_col = max_col or source_ws.max_column
    for row in source_ws.iter_rows(min_row=1, max_row=source_ws.max_row, min_col=1, max_col=max_col):
        for cell in row:
            target_ws.cell(cell.row, cell.column, cell.value)


def day_bounds(row: int) -> tuple[int, int]:
    idx = (row - FIRST_DATA_ROW) // HOURS_PER_DAY
    start = FIRST_DATA_ROW + idx * HOURS_PER_DAY
    return start, start + HOURS_PER_DAY - 1


def create_parameters(ws, battery: float, eta: float, input_data_name: str, core_name: str) -> None:
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Model Parameters and Core Assumptions — Battery {battery:g} kWh"
    ws["A1"].fill = TITLE_FILL; ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    headers = ["Parameter", "Symbol", "Value", "Unit", "Used in", "Notes"]
    for c, value in enumerate(headers, 1):
        cell = ws.cell(2, c, value); cell.fill = HEADER_FILL; cell.font = Font(color="FFFFFF", bold=True); cell.border = BORDER
    for r, row in enumerate(PARAMETER_ROWS, 3):
        vals = list(row)
        if vals[1] == "C_bat": vals[2] = battery; vals[5] = f"Battery-capacity scenario: {battery:g} kWh"
        if vals[1] in ("eta_ch", "eta_dis"): vals[2] = eta
        for c, value in enumerate(vals, 1):
            cell = ws.cell(r, c, value); cell.border = BORDER
            if c == 3: cell.fill = INPUT_FILL
    ws["A25"] = "Standalone builder"; ws["C25"] = "No template workbook required"
    ws["A26"] = "Hourly input"; ws["C26"] = input_data_name
    ws["A27"] = "Core assumptions"; ws["C27"] = core_name
    ws["A29"] = "AK–AL–AM correction"; ws["C29"] = "AK hourly SDR; AL daily charge-energy-weighted price; AM exact EQ13"
    for col, width in {"A":31,"B":18,"C":20,"D":20,"E":30,"F":70}.items(): ws.column_dimensions[col].width = width
    ws.freeze_panes = "A3"


def create_guide(ws, battery: float) -> None:
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Implementation Guide for EQ1–EQ32 — Daily 24-hour optimization — Battery {battery:g} kWh"
    ws["A1"].fill = TITLE_FILL; ws["A1"].font = Font(color="FFFFFF", bold=True, size=14); ws["A1"].alignment = Alignment(horizontal="center")
    headers = ["Equation", "Output parameter", "Unit", "Column", "Excel implementation", "Note"]
    for c, v in enumerate(headers, 1):
        cell=ws.cell(2,c,v); cell.fill=HEADER_FILL; cell.font=Font(color="FFFFFF",bold=True); cell.border=BORDER
    for r,row in enumerate(GUIDE_ROWS,3):
        for c,v in enumerate(row,1):
            cell=ws.cell(r,c,v); cell.border=BORDER; cell.alignment=Alignment(wrap_text=True,vertical="top")
    for col,width in {"A":14,"B":28,"C":17,"D":12,"E":58,"F":35}.items(): ws.column_dimensions[col].width=width
    ws.freeze_panes="A3"


def write_hourly_formulas(ws, pv_area: float) -> None:
    for c,h in enumerate(HEADERS,1):
        cell=ws.cell(1,c,h); cell.fill=HEADER_FILL; cell.font=Font(color="FFFFFF",bold=True); cell.border=BORDER; cell.alignment=Alignment(wrap_text=True,vertical="center")
    ws.row_dimensions[1].height=48
    for r in range(FIRST_DATA_ROW, LAST_DATA_ROW+1):
        ds,de=day_bounds(r)
        # Rebuild PV formula from the core-assumption PV area, preserving 18% efficiency.
        ws[f"H{r}"] = f"=G{r}*{pv_area:g}*0.18"
        formulas = {
            "K":f'=IF(OR(F{r}<=6,F{r}=24),0.55,IF(F{r}<=10,1,IF(F{r}<=16,0.8,IF(F{r}<=22,1.35,0.65))))',
            "L":f'=J{r}/1000', "M":f'=MIN($L${ds}:$L${de})', "N":f'=MAX($L${ds}:$L${de})',
            "O":f'=M{r}+(N{r}-M{r})*Parameters!$C$22', "P":f'=M{r}+(N{r}-M{r})*Parameters!$C$23',
            "Q":f'=MAX($H${ds}:$H${de})',
            "R":f'=IF(I{r}-S{r}-U{r}>0,"Buy from the Grid",IF(AND(U{r}>0,S{r}+U{r}>=I{r}),"discharge",IF(AND(H{r}>=I{r},T{r}>0),"charge",IF(H{r}>S{r},"sell","self consume"))))',
            "S":f'=MAX(0,MIN(I{r},H{r},(K{r}-L{r})/Parameters!$C$3))',
            "T":f'=IF(AND(H{r}>S{r},L{r}<=O{r},V{r}<Parameters!$C$6*Parameters!$C$13),MAX(0,MIN(Parameters!$C$7,H{r}-S{r},(Parameters!$C$6*Parameters!$C$13-V{r})/Parameters!$C$9)),0)',
            "U":f'=IF(AND(I{r}>S{r},L{r}>=P{r},V{r}>Parameters!$C$6*Parameters!$C$12),MAX(0,MIN(Parameters!$C$8,I{r}-S{r},(V{r}-Parameters!$C$6*Parameters!$C$12)*Parameters!$C$10)),0)',
            "V":f'=Parameters!$C$6*Parameters!$C$11' if r==ds else f'=W{r-1}',
            "W":f'=MAX(Parameters!$C$6*Parameters!$C$12,MIN(Parameters!$C$6*Parameters!$C$13,V{r}+Parameters!$C$9*T{r}-U{r}/Parameters!$C$10))',
            "X":f'=MAX(0,I{r}-S{r}-U{r})', "Y":f'=IF(X{r}=0,0,-(1/Parameters!$C$3)*L{r}/X{r})',
            "Z":f'=K{r}*AA{r}-Parameters!$C$3/2*AA{r}^2', "AA":f'=S{r}+U{r}+AG{r}', "AB":f'=I{r}-AA{r}',
            "AC":f'=Parameters!$C$4/2*AB{r}^2', "AD":f'=T{r}-U{r}', "AE":f'=Parameters!$C$5*ABS(AD{r})',
            "AF":f'=IF(R{r}="charge",0,MAX(0,H{r}-S{r}))', "AG":f'=MAX(0,I{r}-S{r}-U{r})',
            "AH":f'=L{r}*(AF{r}-AG{r})', "AI":f'=Z{r}-AC{r}+AH{r}-AE{r}', "AJ":f'=SUM($AI${ds}:$AI${de})',
            "AK":f'=IF(OR(R{r}="charge",I{r}<=0,AN{r}<=0),0,MIN(1,AN{r}/I{r}))',
            "AL":f'=IF(OR(R{r}="charge",AN{r}<=0),0,IFERROR(SUMPRODUCT(($R${ds}:$R${de}="charge")*$T${ds}:$T${de}*$L${ds}:$L${de})/SUMPRODUCT(($R${ds}:$R${de}="charge")*$T${ds}:$T${de}),0))',
            "AM":f'=IF(OR(R{r}="charge",AK{r}=0,AL{r}=0),L{r},IFERROR((L{r}*AL{r})/(((L{r}-AL{r})*AK{r})+AL{r}),L{r}))',
            "AN":f'=IF(R{r}="charge",0,AF{r})', "AO":f'=Z{r}-AC{r}+AN{r}*AM{r}-AE{r}',
            "AP":f'=Parameters!$C$20+(Parameters!$C$21-Parameters!$C$20)*IF(MAX($I${ds}:$I${de})=MIN($I${ds}:$I${de}),0,(I{r}-MIN($I${ds}:$I${de}))/(MAX($I${ds}:$I${de})-MIN($I${ds}:$I${de})))',
            "AQ":f'=IF(MAX($I${ds}:$I${de})=0,0,I{r}/MAX($I${ds}:$I${de}))', "AR":f'=Parameters!$C$15*AP{r}',
            "AS":f'=Parameters!$C$15+AR{r}*AQ{r}^2', "AT":f'=Parameters!$C$16*Parameters!$C$17*AN{r}',
            "AU":f'=Z{r}-AC{r}-AE{r}+AT{r}-AS{r}*AN{r}',
            "AV":f'=K{r}*AN{r}-Parameters!$C$3/2*AN{r}^2-AM{r}*AN{r}+Parameters!$C$16*Parameters!$C$17*AN{r}',
            "AW":f'=K{r}*AN{r}-Parameters!$C$3/2*AN{r}^2+Parameters!$C$16*Parameters!$C$17*AN{r}-K{r}*AA{r}+Parameters!$C$3/2*AA{r}^2+AC{r}+AE{r}',
            "AX":f'=IF(AN{r}=0,"",AW{r}/(2*AN{r}))',
            "AY":f'=(K{r}+Parameters!$C$16*Parameters!$C$17-AM{r})*AN{r}-Parameters!$C$3/2*AN{r}^2',
            "AZ":f'=AN{r}*(AM{r}-Parameters!$C$18-Parameters!$C$5-Parameters!$C$4*ABS(AB{r}))-Parameters!$C$19/2*AN{r}^2',
            "BA":f'=(K{r}+Parameters!$C$16*Parameters!$C$17-Parameters!$C$18-Parameters!$C$5-Parameters!$C$4*ABS(AB{r}))*AN{r}',
            "BB":f'=(Parameters!$C$3+Parameters!$C$19)/2*AN{r}^2',
            "BC":f'=(K{r}+Parameters!$C$16*Parameters!$C$17-Parameters!$C$18-Parameters!$C$5-Parameters!$C$4*ABS(AB{r}))*AN{r}-((Parameters!$C$3+Parameters!$C$19)/2)*AN{r}^2',
            "BD":f'=IF(R{r}="charge",X{r}*(K{r}+Parameters!$C$3*Parameters!$C$7-Parameters!$C$3*H{r}-L{r})-Parameters!$C$3/2*X{r}^2,IF(R{r}="discharge",X{r}*(K{r}-Parameters!$C$3*Parameters!$C$8-Parameters!$C$3*H{r}-L{r})-Parameters!$C$3/2*X{r}^2,X{r}*(K{r}-Parameters!$C$3*H{r}-L{r})-Parameters!$C$3/2*X{r}^2))',
            "BF":f'=MAX($L${ds}:$L${de})', "BG":f'=MAX(0,(BF{r}-L{r})*U{r})', "BH":f'=BD{r}+BG{r}'
        }
        for col,formula in formulas.items(): ws[f"{col}{r}"]=formula
        ws[f"BE{r}"] = None
    # formats and dimensions
    ws.freeze_panes="AQ7"; ws.auto_filter.ref=f"A1:BH{LAST_DATA_ROW}"
    widths={"A":20,"B":12,"C":12,"D":12,"E":8,"F":8,"G":18,"H":20,"I":20,"J":18,"K":14,"L":16,"R":19,"AK":18,"AL":24,"AM":18}
    for c in range(1,61): ws.column_dimensions[get_column_letter(c)].width=widths.get(get_column_letter(c),16)
    # Keep formatting lightweight so generation remains practical in Colab.
    # Excel will display formulas normally; headers and worksheet layout are styled.
    for r in range(2, LAST_DATA_ROW + 1):
        ws[f"A{r}"].number_format = "yyyy-mm-dd hh:mm"
        ws[f"B{r}"].number_format = "yyyy-mm-dd"


def create_summary(ws, battery: float) -> None:
    ws.merge_cells("A1:L1")
    ws["A1"]=f"Daily 24-hour Utility and Welfare Aggregations — Battery {battery:g} kWh"
    ws["A1"].fill=TITLE_FILL; ws["A1"].font=Font(color="FFFFFF",bold=True,size=14); ws["A1"].alignment=Alignment(horizontal="center")
    headers=["Date","Daily Umax EQ8/29","Daily Uprosumer EQ14","Daily SWtotal EQ15","Daily CS EQ23","Daily PS EQ24","Daily SWP2P EQ25","Daily CSpros EQ28","Daily SWgrid EQ32",None,"Annual KPI","Value (€)"]
    for c,v in enumerate(headers,1):
        cell=ws.cell(2,c,v); cell.fill=HEADER_FILL; cell.font=Font(color="FFFFFF",bold=True); cell.border=BORDER
    source_cols=["AI","AO","AU","AY","AZ","BC","BD","BH"]
    for day in range(365):
        out_r=3+day; start=2+day*24; end=start+23
        ws.cell(out_r,1,f"={HOURLY_SHEET}!B{start}")
        for j,col in enumerate(source_cols,2): ws.cell(out_r,j,f"=SUM({HOURLY_SHEET}!{col}{start}:{col}{end})")
        for c in range(1,10): ws.cell(out_r,c).border=BORDER; ws.cell(out_r,c).number_format="0.000000"
        ws.cell(out_r,1).number_format="yyyy-mm-dd"
    kpis=[("Annual Umax EQ8/29","B"),("Annual Uprosumer EQ14","C"),("Annual SWtotal EQ15","D"),("Annual CS EQ23","E"),("Annual PS EQ24","F"),("Annual SWP2P EQ25","G"),("Annual CSpros EQ28","H"),("Annual SWgrid EQ32","I")]
    for i,(label,col) in enumerate(kpis,3): ws.cell(i,11,label); ws.cell(i,12,f"=SUM({col}3:{col}367)")
    for col,width in {"A":14,"B":20,"C":22,"D":20,"E":18,"F":18,"G":20,"H":20,"I":20,"J":3,"K":24,"L":18}.items(): ws.column_dimensions[col].width=width
    ws.freeze_panes="A3"


def create_workbook(input_data: Path, core: Path, battery: float, pv_area: float, eta: float) -> Workbook:
    source = load_workbook(input_data, data_only=False, read_only=False)
    if MONTHLY_SHEET not in source.sheetnames or HOURLY_SHEET not in source.sheetnames:
        raise KeyError(f"Input Data must contain sheets {MONTHLY_SHEET!r} and {HOURLY_SHEET!r}")
    if source[HOURLY_SHEET].max_row < LAST_DATA_ROW or source[HOURLY_SHEET].max_column < 10:
        raise ValueError("Input hourly sheet must contain at least 8760 data rows and columns A:J")
    wb=Workbook(); wb.remove(wb.active)
    monthly=wb.create_sheet(MONTHLY_SHEET); hourly=wb.create_sheet(HOURLY_SHEET); params=wb.create_sheet(PARAMETERS_SHEET); guide=wb.create_sheet(GUIDE_SHEET); summary=wb.create_sheet(SUMMARY_SHEET)
    copy_input_sheet_values(source[MONTHLY_SHEET],monthly)
    copy_input_sheet_values(source[HOURLY_SHEET],hourly,max_col=10)
    # Basic monthly formatting
    monthly.freeze_panes="A2"; monthly.auto_filter.ref=f"A1:H{monthly.max_row}"
    for c in range(1,monthly.max_column+1): monthly.column_dimensions[get_column_letter(c)].width=22
    for c in range(1,monthly.max_column+1):
        monthly.cell(1,c).fill=HEADER_FILL; monthly.cell(1,c).font=Font(color="FFFFFF",bold=True)
    create_parameters(params,battery,eta,input_data.name,core.name)
    create_guide(guide,battery)
    write_hourly_formulas(hourly,pv_area)
    create_summary(summary,battery)
    wb.calculation.fullCalcOnLoad=True; wb.calculation.forceFullCalc=True; wb.calculation.calcMode="auto"
    wb.active=1
    return wb


def verify_workbook(path: Path, battery: float) -> None:
    wb=load_workbook(path,data_only=False,read_only=True)
    required=[MONTHLY_SHEET,HOURLY_SHEET,PARAMETERS_SHEET,GUIDE_SHEET,SUMMARY_SHEET]
    for s in required:
        if s not in wb.sheetnames: raise AssertionError(f"Missing sheet: {s}")
    ws=wb[HOURLY_SHEET]
    checks={"Z2":"=K2*AA2-Parameters!$C$3/2*AA2^2","AZ2":"=AN2*(AM2-Parameters!$C$18-Parameters!$C$5-Parameters!$C$4*ABS(AB2))-Parameters!$C$19/2*AN2^2"}
    for cell,expected in checks.items():
        if ws[cell].value != expected: raise AssertionError(f"Formula check failed in {cell}")
    if ws["BE2"].value is not None: raise AssertionError("BE must be blank")
    if float(wb[PARAMETERS_SHEET]["C6"].value)!=float(battery): raise AssertionError("Battery scenario mismatch")


def update_scenario_in_existing_workbook(path: Path, battery: float) -> None:
    wb = load_workbook(path, data_only=False, read_only=False)
    wb[PARAMETERS_SHEET]["C6"] = float(battery)
    wb[PARAMETERS_SHEET]["F6"] = f"Battery-capacity scenario: {battery:g} kWh"
    wb[PARAMETERS_SHEET]["A1"] = f"Model Parameters and Core Assumptions — Battery {battery:g} kWh"
    wb[GUIDE_SHEET]["A1"] = f"Implementation Guide for EQ1–EQ32 — Daily 24-hour optimization — Battery {battery:g} kWh"
    wb[SUMMARY_SHEET]["A1"] = f"Daily 24-hour Utility and Welfare Aggregations — Battery {battery:g} kWh"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(path)


def main() -> int:
    args=parse_args()
    try:
        input_data=existing_or_fallback(args.input_data,["Input Data*.xlsx"])
        core=existing_or_fallback(args.core_assumptions,["Input Core assumptions*.xlsx"])
        pv_area,_,eta=read_core_assumptions(core)
        args.output_dir.mkdir(parents=True,exist_ok=True)
        sizes=[float(x) for x in args.battery_sizes]
        if not sizes:
            raise ValueError("At least one battery size is required")

        # Build the complete workbook only once from the two inputs. The remaining
        # battery cases are byte-level copies whose scenario parameter/title cells
        # are updated. No external template workbook is ever read.
        first=sizes[0]
        first_label=f"{first:g}".replace(".","p")
        first_out=args.output_dir/f"Output_EQ1_EQ32_DailyOptimization_Battery_{first_label}kWh.xlsx"
        wb=create_workbook(input_data,core,first,pv_area,eta)
        wb.save(first_out)
        verify_workbook(first_out,first)
        outputs=[first_out]
        print(f"Created: {first_out}")

        import shutil
        for battery in sizes[1:]:
            label=f"{battery:g}".replace(".","p")
            out=args.output_dir/f"Output_EQ1_EQ32_DailyOptimization_Battery_{label}kWh.xlsx"
            shutil.copy2(first_out,out)
            update_scenario_in_existing_workbook(out,battery)
            verify_workbook(out,battery)
            outputs.append(out)
            print(f"Created: {out}")

        with zipfile.ZipFile(args.zip_output,"w",compression=zipfile.ZIP_DEFLATED) as zf:
            for out in outputs: zf.write(out,arcname=out.name)
        print(f"Created ZIP: {args.zip_output}")
        return 0
    except Exception as exc:
        print(f"ERROR: {exc}",file=sys.stderr)
        return 1



# ============================================================================
# MULTI-STRATEGY EXTENSION (v3)
# Strategy A: baseline fixed daily thresholds
# Strategy B: adaptive 24-hour forecast rules, grid charging, storage-first sales,
#             dynamic SOC reserve
# Strategy C: rolling 24-hour cost/welfare optimization using scipy.optimize.linprog
# ============================================================================

import math
from copy import copy

try:
    import numpy as np
    from scipy.optimize import linprog
except Exception:  # Strategy A/B can still be generated without scipy
    np = None
    linprog = None

STRATEGIES = ("A", "B", "C")
STRATEGY_NAMES = {
    "A": "Baseline fixed-threshold dispatch",
    "B": "Adaptive forecast-based dispatch",
    "C": "Rolling 24-hour optimized dispatch",
}


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Create EQ1-EQ32 workbooks for 3 dispatch strategies and multiple battery sizes."
    )
    p.add_argument("--input-data", type=Path, default=Path("Input Data.xlsx"))
    p.add_argument("--core-assumptions", type=Path, default=Path("Input Core assumptions.xlsx"))
    p.add_argument("--output-dir", type=Path, default=Path("outputs_EQ1_EQ32_multistrategy"))
    p.add_argument("--zip-output", type=Path, default=Path("Output_EQ1_EQ32_3Strategies_4Batteries.zip"))
    p.add_argument("--battery-sizes", type=float, nargs="+", default=list(BATTERY_SIZES))
    p.add_argument("--strategies", nargs="+", choices=list(STRATEGIES), default=list(STRATEGIES))
    p.add_argument("--grid-charge-limit", type=float, default=3.0, help="Maximum grid charging energy per hour (kWh).")
    p.add_argument("--sell-discharge-limit", type=float, default=3.0, help="Maximum battery export per hour (kWh).")
    p.add_argument("--peak-penalty", type=float, default=0.04, help="Penalty weight for grid imports during high-price hours (EUR/kWh proxy).")
    p.add_argument("--self-consumption-reward", type=float, default=0.02, help="Reward for charging PV surplus (EUR/kWh proxy).")
    p.add_argument("--terminal-soc", choices=["daily", "annual", "none"], default="daily")
    args, unknown = p.parse_known_args()
    return args


def add_strategy_parameters(ws, strategy: str, args: argparse.Namespace) -> None:
    start = 31
    rows = [
        ("Dispatch strategy", "strategy", STRATEGY_NAMES[strategy], "text"),
        ("Grid charging limit", "P_grid_ch_max", args.grid_charge_limit, "kWh/h"),
        ("Battery-to-market export limit", "P_sell_dis_max", args.sell_discharge_limit, "kWh/h"),
        ("Peak import penalty", "lambda_peak", args.peak_penalty, "EUR/kWh proxy"),
        ("PV self-consumption reward", "lambda_self", args.self_consumption_reward, "EUR/kWh proxy"),
        ("Terminal SOC condition", "terminal_soc", args.terminal_soc, "text"),
        ("Adaptive reserve: base", "reserve_base", 0.10, "fraction"),
        ("Adaptive reserve: evening", "reserve_evening", 0.30, "fraction"),
        ("Adaptive low-price share", "adaptive_low", 0.30, "fraction"),
        ("Adaptive high-price share", "adaptive_high", 0.70, "fraction"),
        ("Arbitrage safety margin", "arb_margin", 0.01, "EUR/kWh"),
        ("Hourly self-discharge", "sigma_h", 0.00002, "fraction/hour"),
    ]
    ws[f"A{start-1}"] = "Strategy-specific assumptions"
    ws[f"A{start-1}"].fill = SECTION_FILL
    ws[f"A{start-1}"].font = Font(bold=True)
    for i, row in enumerate(rows, start):
        for c, value in enumerate(row, 1):
            ws.cell(i, c, value)
            ws.cell(i, c).border = BORDER
        ws.cell(i, 3).fill = INPUT_FILL


def create_parameters(ws, battery: float, eta: float, input_data_name: str, core_name: str,
                      strategy: str = "A", args: argparse.Namespace | None = None) -> None:
    # Use original builder's parameter layout.
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Model Parameters — Battery {battery:g} kWh — Strategy {strategy}"
    ws["A1"].fill = TITLE_FILL; ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    headers = ["Parameter", "Symbol", "Value", "Unit", "Used in", "Notes"]
    for c, value in enumerate(headers, 1):
        cell = ws.cell(2, c, value); cell.fill = HEADER_FILL; cell.font = Font(color="FFFFFF", bold=True); cell.border = BORDER
    for r, row in enumerate(PARAMETER_ROWS, 3):
        vals = list(row)
        if vals[1] == "C_bat": vals[2] = battery; vals[5] = f"Battery-capacity scenario: {battery:g} kWh"
        if vals[1] in ("eta_ch", "eta_dis"): vals[2] = eta
        for c, value in enumerate(vals, 1):
            cell = ws.cell(r, c, value); cell.border = BORDER
            if c == 3: cell.fill = INPUT_FILL
    ws["A25"] = "Standalone builder"; ws["C25"] = "No template workbook required"
    ws["A26"] = "Hourly input"; ws["C26"] = input_data_name
    ws["A27"] = "Core assumptions"; ws["C27"] = core_name
    ws["A29"] = "Dispatch strategy"; ws["C29"] = STRATEGY_NAMES[strategy]
    if args is not None:
        add_strategy_parameters(ws, strategy, args)
    for col, width in {"A":31,"B":18,"C":24,"D":20,"E":30,"F":70}.items(): ws.column_dimensions[col].width = width
    ws.freeze_panes = "A3"


def _future_end(r: int) -> int:
    return min(LAST_DATA_ROW, r + 23)


def write_hourly_formulas_strategy_ab(ws, pv_area: float, strategy: str) -> None:
    """Write formula-driven Strategy A or B workbooks."""
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, c, h); cell.fill = HEADER_FILL; cell.font = Font(color="FFFFFF", bold=True)
        cell.border = BORDER; cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 48

    for r in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
        ds, de = day_bounds(r)
        fe = _future_end(r)
        ws[f"H{r}"] = f"=G{r}*{pv_area:g}*0.18"
        common = {
            "K":f'=IF(OR(F{r}<=6,F{r}=24),0.55,IF(F{r}<=10,1,IF(F{r}<=16,0.8,IF(F{r}<=22,1.35,0.65))))',
            "L":f'=J{r}/1000', "M":f'=MIN($L${ds}:$L${de})', "N":f'=MAX($L${ds}:$L${de})',
            "O":f'=M{r}+(N{r}-M{r})*Parameters!$C$22', "P":f'=M{r}+(N{r}-M{r})*Parameters!$C$23',
            "Q":f'=MAX($H${ds}:$H${de})',
            "S":f'=MAX(0,MIN(I{r},H{r},(K{r}-L{r})/Parameters!$C$3))',
        }
        if strategy == "A":
            t_formula = f'=IF(AND(H{r}>S{r},L{r}<=O{r},V{r}<Parameters!$C$6*Parameters!$C$13),MAX(0,MIN(Parameters!$C$7,H{r}-S{r},(Parameters!$C$6*Parameters!$C$13-V{r})/Parameters!$C$9)),0)'
            u_formula = f'=IF(AND(I{r}>S{r},L{r}>=P{r},V{r}>Parameters!$C$6*Parameters!$C$12),MAX(0,MIN(Parameters!$C$8,I{r}-S{r},(V{r}-Parameters!$C$6*Parameters!$C$12)*Parameters!$C$10)),0)'
        else:
            # Strategy B: adaptive rolling 24h thresholds; PV-first charge, optional grid charge,
            # dynamic reserve, and battery export at sufficiently high prices.
            fmin = f'MIN($L${r}:$L${fe})'
            fmax = f'MAX($L${r}:$L${fe})'
            low = f'({fmin}+({fmax}-{fmin})*Parameters!$C$39)'
            high = f'({fmin}+({fmax}-{fmin})*Parameters!$C$40)'
            reserve = f'IF(AND(F{r}>=17,F{r}<=22),Parameters!$C$38,Parameters!$C$37)'
            capacity_room = f'(Parameters!$C$6*Parameters!$C$13-V{r})/Parameters!$C$9'
            pv_surplus = f'MAX(0,H{r}-S{r})'
            future_arbitrage = f'({fmax}*Parameters!$C$9*Parameters!$C$10>L{r}+Parameters!$C$41+Parameters!$C$5)'
            # Charge PV first. Grid charge is allowed only at the adaptive low threshold and
            # only where future arbitrage remains profitable after losses and degradation.
            t_formula = (
                f'=IF(V{r}>=Parameters!$C$6*Parameters!$C$13,0,'
                f'MAX(0,MIN(Parameters!$C$7,{capacity_room},'
                f'IF({pv_surplus}>0,IF(OR(L{r}<={low},{future_arbitrage}),{pv_surplus},0),'
                f'IF(AND(L{r}<={low},{future_arbitrage}),Parameters!$C$32,0)))))'
            )
            residual = f'MAX(0,I{r}-S{r})'
            export_room = f'IF(L{r}>={high},Parameters!$C$33,0)'
            u_formula = (
                f'=IF(AND(L{r}>={high},V{r}>Parameters!$C$6*{reserve}),'
                f'MAX(0,MIN(Parameters!$C$8,{residual}+{export_room},'
                f'(V{r}-Parameters!$C$6*{reserve})*Parameters!$C$10)),0)'
            )

        formulas = dict(common)
        formulas.update({
            "T": t_formula, "U": u_formula,
            "V": f'=Parameters!$C$6*Parameters!$C$11' if r == ds else f'=W{r-1}',
            "W": f'=MAX(Parameters!$C$6*Parameters!$C$12,MIN(Parameters!$C$6*Parameters!$C$13,V{r}*(1-Parameters!$C$42)+Parameters!$C$9*T{r}-U{r}/Parameters!$C$10))',
            "X":f'=MAX(0,I{r}-S{r}-MIN(U{r},MAX(0,I{r}-S{r})))',
            "Y":f'=IF(X{r}=0,0,-(1/Parameters!$C$3)*L{r}/X{r})',
            "Z":f'=K{r}*AA{r}-Parameters!$C$3/2*AA{r}^2',
            "AA":f'=S{r}+MIN(U{r},MAX(0,I{r}-S{r}))+AG{r}', "AB":f'=I{r}-AA{r}',
            "AC":f'=Parameters!$C$4/2*AB{r}^2', "AD":f'=T{r}-U{r}', "AE":f'=Parameters!$C$5*ABS(AD{r})',
            # Charge uses PV first; any excess charge is grid charging. Discharge serves load first,
            # then can be exported.
            "AF":f'=MAX(0,H{r}-S{r}-MIN(T{r},MAX(0,H{r}-S{r}))+MAX(0,U{r}-MAX(0,I{r}-S{r})))',
            "AG":f'=MAX(0,I{r}-S{r}-MIN(U{r},MAX(0,I{r}-S{r})))+MAX(0,T{r}-MAX(0,H{r}-S{r}))',
            "AH":f'=L{r}*(AF{r}-AG{r})', "AI":f'=Z{r}-AC{r}+AH{r}-AE{r}', "AJ":f'=SUM($AI${ds}:$AI${de})',
            "AK":f'=IF(OR(I{r}<=0,AN{r}<=0),0,MIN(1,AN{r}/I{r}))',
            "AL":f'=IF(AN{r}<=0,0,IFERROR(SUMPRODUCT($T${ds}:$T${de},$L${ds}:$L${de})/SUM($T${ds}:$T${de}),0))',
            "AM":f'=IF(OR(AK{r}=0,AL{r}=0),L{r},IFERROR((L{r}*AL{r})/(((L{r}-AL{r})*AK{r})+AL{r}),L{r}))',
            "AN":f'=AF{r}', "AO":f'=Z{r}-AC{r}+AN{r}*AM{r}-AE{r}',
            "AP":f'=Parameters!$C$20+(Parameters!$C$21-Parameters!$C$20)*IF(MAX($I${ds}:$I${de})=MIN($I${ds}:$I${de}),0,(I{r}-MIN($I${ds}:$I${de}))/(MAX($I${ds}:$I${de})-MIN($I${ds}:$I${de})))',
            "AQ":f'=IF(MAX($I${ds}:$I${de})=0,0,I{r}/MAX($I${ds}:$I${de}))', "AR":f'=Parameters!$C$15*AP{r}',
            "AS":f'=Parameters!$C$15+AR{r}*AQ{r}^2', "AT":f'=Parameters!$C$16*Parameters!$C$17*AN{r}',
            "AU":f'=Z{r}-AC{r}-AE{r}+AT{r}-AS{r}*AN{r}',
            "AV":f'=K{r}*AN{r}-Parameters!$C$3/2*AN{r}^2-AM{r}*AN{r}+Parameters!$C$16*Parameters!$C$17*AN{r}',
            "AW":f'=K{r}*AN{r}-Parameters!$C$3/2*AN{r}^2+Parameters!$C$16*Parameters!$C$17*AN{r}-K{r}*AA{r}+Parameters!$C$3/2*AA{r}^2+AC{r}+AE{r}',
            "AX":f'=IF(AN{r}=0,"",AW{r}/(2*AN{r}))',
            "AY":f'=(K{r}+Parameters!$C$16*Parameters!$C$17-AM{r})*AN{r}-Parameters!$C$3/2*AN{r}^2',
            "AZ":f'=AN{r}*(AM{r}-Parameters!$C$18-Parameters!$C$5-Parameters!$C$4*ABS(AB{r}))-Parameters!$C$19/2*AN{r}^2',
            "BA":f'=(K{r}+Parameters!$C$16*Parameters!$C$17-Parameters!$C$18-Parameters!$C$5-Parameters!$C$4*ABS(AB{r}))*AN{r}',
            "BB":f'=(Parameters!$C$3+Parameters!$C$19)/2*AN{r}^2',
            "BC":f'=(K{r}+Parameters!$C$16*Parameters!$C$17-Parameters!$C$18-Parameters!$C$5-Parameters!$C$4*ABS(AB{r}))*AN{r}-((Parameters!$C$3+Parameters!$C$19)/2)*AN{r}^2',
            "BD":f'=X{r}*(K{r}-Parameters!$C$3*H{r}-L{r})-Parameters!$C$3/2*X{r}^2',
            "BF":f'=MAX($L${ds}:$L${de})', "BG":f'=MAX(0,(BF{r}-L{r})*MIN(U{r},MAX(0,I{r}-S{r})))', "BH":f'=BD{r}+BG{r}',
            "R":f'=IF(AG{r}>MAX(0,I{r}-S{r}-U{r}),"Grid charge",IF(U{r}>MAX(0,I{r}-S{r}),"Battery sell",IF(U{r}>0,"discharge",IF(T{r}>0,"charge",IF(AF{r}>0,"sell",IF(S{r}>0,"self consume","Buy from the Grid"))))))'
        })
        for col, formula in formulas.items(): ws[f"{col}{r}"] = formula
        ws[f"BE{r}"] = None

    ws.freeze_panes = "AQ7"; ws.auto_filter.ref = f"A1:BH{LAST_DATA_ROW}"
    widths = {"A":20,"B":12,"C":12,"D":12,"E":8,"F":8,"G":18,"H":20,"I":20,"J":18,"K":14,"L":16,"R":19,"AK":18,"AL":24,"AM":18}
    for c in range(1, 61): ws.column_dimensions[get_column_letter(c)].width = widths.get(get_column_letter(c), 16)
    for rr in range(2, LAST_DATA_ROW + 1):
        ws[f"A{rr}"].number_format = "yyyy-mm-dd hh:mm"; ws[f"B{rr}"].number_format = "yyyy-mm-dd"


def _num(v, default=0.0):
    try:
        return float(v)
    except Exception:
        return default


def optimize_day_lp(load, pv, price, battery, eta_ch, eta_dis, soc0, p_ch, p_dis,
                    grid_ch_limit, sell_dis_limit, psi, peak_penalty, self_reward,
                    reserve_frac, terminal_equal=True):
    """Continuous 24-hour rolling optimization.

    Variables per hour: pv_charge, grid_charge, discharge_load, discharge_sell, SOC_end.
    The LP minimizes purchase cost minus export revenue, degradation cost and peak penalty,
    while rewarding PV storage. Direct PV self-consumption is fixed before optimization.
    """
    if linprog is None or np is None:
        raise RuntimeError("Strategy C requires scipy and numpy: pip install scipy numpy")
    n = len(load)
    direct = np.minimum(load, pv)
    residual = np.maximum(0.0, load - direct)
    surplus = np.maximum(0.0, pv - direct)
    # Conservative market export price: 80% of spot plus a small green premium.
    sell_price = np.minimum(price, 0.80 * price + 0.022)
    peak_cut = np.quantile(price, 0.70) if n else 0.0

    # indices
    def ix(block, t): return block*n + t
    N = 5*n
    c = np.zeros(N)
    for t in range(n):
        high = 1.0 if price[t] >= peak_cut else 0.0
        c[ix(0,t)] = psi - self_reward                  # PV charging
        c[ix(1,t)] = price[t] + psi                     # grid charging
        c[ix(2,t)] = -price[t] + psi - peak_penalty*high # discharge to load
        c[ix(3,t)] = -sell_price[t] + psi               # discharge to market
        c[ix(4,t)] = 0.0

    bounds = []
    for t in range(n): bounds.append((0, min(p_ch, surplus[t])))
    for t in range(n): bounds.append((0, min(p_ch, grid_ch_limit)))
    for t in range(n): bounds.append((0, min(p_dis, residual[t])))
    for t in range(n): bounds.append((0, min(p_dis, sell_dis_limit)))
    for t in range(n): bounds.append((battery*reserve_frac[t], battery))

    A_eq=[]; b_eq=[]
    sigma=0.00002
    for t in range(n):
        row=np.zeros(N)
        row[ix(4,t)] = 1
        row[ix(0,t)] = -eta_ch
        row[ix(1,t)] = -eta_ch
        row[ix(2,t)] = 1/eta_dis
        row[ix(3,t)] = 1/eta_dis
        if t == 0:
            b = soc0*(1-sigma)
        else:
            row[ix(4,t-1)] = -(1-sigma)
            b = 0
        A_eq.append(row); b_eq.append(b)
    if terminal_equal:
        row=np.zeros(N); row[ix(4,n-1)]=1
        A_eq.append(row); b_eq.append(soc0)

    # Combined power limits: pv_charge + grid_charge <= p_ch; dis_load + dis_sell <= p_dis
    A_ub=[]; b_ub=[]
    for t in range(n):
        row=np.zeros(N); row[ix(0,t)]=1; row[ix(1,t)]=1
        A_ub.append(row); b_ub.append(p_ch)
        row=np.zeros(N); row[ix(2,t)]=1; row[ix(3,t)]=1
        A_ub.append(row); b_ub.append(p_dis)

    res=linprog(c, A_ub=np.array(A_ub), b_ub=np.array(b_ub),
                A_eq=np.array(A_eq), b_eq=np.array(b_eq), bounds=bounds, method="highs")
    if not res.success:
        raise RuntimeError(f"Rolling LP failed: {res.message}")
    x=res.x
    return {
        "direct": direct,
        "pv_ch": np.array([x[ix(0,t)] for t in range(n)]),
        "grid_ch": np.array([x[ix(1,t)] for t in range(n)]),
        "dis_load": np.array([x[ix(2,t)] for t in range(n)]),
        "dis_sell": np.array([x[ix(3,t)] for t in range(n)]),
        "soc_end": np.array([x[ix(4,t)] for t in range(n)]),
        "residual": residual,
        "surplus": surplus,
    }


def write_hourly_strategy_c(ws, pv_area: float, battery: float, eta: float, args: argparse.Namespace) -> None:
    if linprog is None:
        raise RuntimeError("Strategy C requires scipy: pip install scipy numpy")
    for c, h in enumerate(HEADERS, 1):
        cell=ws.cell(1,c,h); cell.fill=HEADER_FILL; cell.font=Font(color="FFFFFF",bold=True); cell.border=BORDER
        cell.alignment=Alignment(wrap_text=True,vertical="center")
    ws.row_dimensions[1].height=48

    # Build numeric arrays from copied A:J input data.
    load=[]; pv=[]; price=[]
    for r in range(FIRST_DATA_ROW, LAST_DATA_ROW+1):
        g=_num(ws[f"G{r}"].value)
        h=g*pv_area*0.18
        ws[f"H{r}"] = h
        load.append(_num(ws[f"I{r}"].value))
        price.append(_num(ws[f"J{r}"].value)/1000.0)
        pv.append(h)
    load=np.asarray(load); pv=np.asarray(pv); price=np.asarray(price)

    T=np.zeros(8760); U=np.zeros(8760); V=np.zeros(8760); W=np.zeros(8760)
    AF=np.zeros(8760); AG=np.zeros(8760); S=np.minimum(load,pv)
    soc_prev=battery*0.90
    for d in range(365):
        a=d*24; b=a+24
        hrs=np.arange(24)
        # Dynamic reserve: 30% for evening hours and 10% otherwise.
        reserve=np.where((hrs>=16)&(hrs<=21),0.30,0.10)
        result=optimize_day_lp(load[a:b],pv[a:b],price[a:b],battery,eta,eta,soc_prev,
                               3.0,3.0,args.grid_charge_limit,args.sell_discharge_limit,
                               0.03,args.peak_penalty,args.self_consumption_reward,reserve,
                               terminal_equal=(args.terminal_soc=="daily"))
        pv_ch=result["pv_ch"]; grid_ch=result["grid_ch"]
        dis_load=result["dis_load"]; dis_sell=result["dis_sell"]
        T[a:b]=pv_ch+grid_ch; U[a:b]=dis_load+dis_sell
        V[a]=soc_prev
        for t in range(24):
            if t>0: V[a+t]=W[a+t-1]
            W[a+t]=result["soc_end"][t]
        AF[a:b]=result["surplus"]-pv_ch+dis_sell
        AG[a:b]=result["residual"]-dis_load+grid_ch
        soc_prev=W[b-1]

    for idx,r in enumerate(range(FIRST_DATA_ROW,LAST_DATA_ROW+1)):
        ds,de=day_bounds(r)
        ws[f"K{r}"]=f'=IF(OR(F{r}<=6,F{r}=24),0.55,IF(F{r}<=10,1,IF(F{r}<=16,0.8,IF(F{r}<=22,1.35,0.65))))'
        ws[f"L{r}"]=f'=J{r}/1000'; ws[f"M{r}"]=f'=MIN($L${ds}:$L${de})'; ws[f"N{r}"]=f'=MAX($L${ds}:$L${de})'
        ws[f"O{r}"]=f'=M{r}+(N{r}-M{r})*Parameters!$C$22'; ws[f"P{r}"]=f'=M{r}+(N{r}-M{r})*Parameters!$C$23'; ws[f"Q{r}"]=f'=MAX($H${ds}:$H${de})'
        ws[f"S{r}"]=float(S[idx]); ws[f"T{r}"]=float(T[idx]); ws[f"U{r}"]=float(U[idx]); ws[f"V{r}"]=float(V[idx]); ws[f"W{r}"]=float(W[idx])
        ws[f"AF{r}"]=float(max(0,AF[idx])); ws[f"AG{r}"]=float(max(0,AG[idx]))
        ws[f"R{r}"] = ("Grid charge" if T[idx] > max(0,pv[idx]-S[idx])+1e-8 else
                         "Battery sell" if U[idx] > max(0,load[idx]-S[idx])+1e-8 else
                         "discharge" if U[idx]>1e-8 else "charge" if T[idx]>1e-8 else
                         "sell" if AF[idx]>1e-8 else "self consume" if S[idx]>1e-8 else "Buy from the Grid")
        formulas={
            "X":f'=MAX(0,I{r}-S{r}-MIN(U{r},MAX(0,I{r}-S{r})))', "Y":f'=IF(X{r}=0,0,-(1/Parameters!$C$3)*L{r}/X{r})',
            "Z":f'=K{r}*AA{r}-Parameters!$C$3/2*AA{r}^2', "AA":f'=S{r}+MIN(U{r},MAX(0,I{r}-S{r}))+AG{r}', "AB":f'=I{r}-AA{r}',
            "AC":f'=Parameters!$C$4/2*AB{r}^2', "AD":f'=T{r}-U{r}', "AE":f'=Parameters!$C$5*ABS(AD{r})',
            "AH":f'=L{r}*(AF{r}-AG{r})', "AI":f'=Z{r}-AC{r}+AH{r}-AE{r}', "AJ":f'=SUM($AI${ds}:$AI${de})',
            "AK":f'=IF(OR(I{r}<=0,AN{r}<=0),0,MIN(1,AN{r}/I{r}))',
            "AL":f'=IF(AN{r}<=0,0,IFERROR(SUMPRODUCT($T${ds}:$T${de},$L${ds}:$L${de})/SUM($T${ds}:$T${de}),0))',
            "AM":f'=IF(OR(AK{r}=0,AL{r}=0),L{r},IFERROR((L{r}*AL{r})/(((L{r}-AL{r})*AK{r})+AL{r}),L{r}))',
            "AN":f'=AF{r}', "AO":f'=Z{r}-AC{r}+AN{r}*AM{r}-AE{r}',
            "AP":f'=Parameters!$C$20+(Parameters!$C$21-Parameters!$C$20)*IF(MAX($I${ds}:$I${de})=MIN($I${ds}:$I${de}),0,(I{r}-MIN($I${ds}:$I${de}))/(MAX($I${ds}:$I${de})-MIN($I${ds}:$I${de})))',
            "AQ":f'=IF(MAX($I${ds}:$I${de})=0,0,I{r}/MAX($I${ds}:$I${de}))', "AR":f'=Parameters!$C$15*AP{r}', "AS":f'=Parameters!$C$15+AR{r}*AQ{r}^2',
            "AT":f'=Parameters!$C$16*Parameters!$C$17*AN{r}', "AU":f'=Z{r}-AC{r}-AE{r}+AT{r}-AS{r}*AN{r}',
            "AV":f'=K{r}*AN{r}-Parameters!$C$3/2*AN{r}^2-AM{r}*AN{r}+Parameters!$C$16*Parameters!$C$17*AN{r}',
            "AW":f'=K{r}*AN{r}-Parameters!$C$3/2*AN{r}^2+Parameters!$C$16*Parameters!$C$17*AN{r}-K{r}*AA{r}+Parameters!$C$3/2*AA{r}^2+AC{r}+AE{r}',
            "AX":f'=IF(AN{r}=0,"",AW{r}/(2*AN{r}))', "AY":f'=(K{r}+Parameters!$C$16*Parameters!$C$17-AM{r})*AN{r}-Parameters!$C$3/2*AN{r}^2',
            "AZ":f'=AN{r}*(AM{r}-Parameters!$C$18-Parameters!$C$5-Parameters!$C$4*ABS(AB{r}))-Parameters!$C$19/2*AN{r}^2',
            "BA":f'=(K{r}+Parameters!$C$16*Parameters!$C$17-Parameters!$C$18-Parameters!$C$5-Parameters!$C$4*ABS(AB{r}))*AN{r}',
            "BB":f'=(Parameters!$C$3+Parameters!$C$19)/2*AN{r}^2',
            "BC":f'=(K{r}+Parameters!$C$16*Parameters!$C$17-Parameters!$C$18-Parameters!$C$5-Parameters!$C$4*ABS(AB{r}))*AN{r}-((Parameters!$C$3+Parameters!$C$19)/2)*AN{r}^2',
            "BD":f'=X{r}*(K{r}-Parameters!$C$3*H{r}-L{r})-Parameters!$C$3/2*X{r}^2',
            "BF":f'=MAX($L${ds}:$L${de})', "BG":f'=MAX(0,(BF{r}-L{r})*MIN(U{r},MAX(0,I{r}-S{r})))', "BH":f'=BD{r}+BG{r}'
        }
        for col,formula in formulas.items(): ws[f"{col}{r}"]=formula
        ws[f"BE{r}"]=None
    ws.freeze_panes="AQ7"; ws.auto_filter.ref=f"A1:BH{LAST_DATA_ROW}"
    for c in range(1,61): ws.column_dimensions[get_column_letter(c)].width=16
    ws.column_dimensions["A"].width=20; ws.column_dimensions["R"].width=19


def create_summary(ws, battery: float, strategy: str = "A") -> None:
    ws.merge_cells("A1:L1")
    ws["A1"] = f"Annual and Daily KPIs — Battery {battery:g} kWh — Strategy {strategy}"
    ws["A1"].fill=TITLE_FILL; ws["A1"].font=Font(color="FFFFFF",bold=True,size=14); ws["A1"].alignment=Alignment(horizontal="center")
    headers=["Date","Daily Umax","Daily Uprosumer","Daily SWtotal","Daily CS","Daily PS","Daily SWP2P","Daily CSpros","Daily SWgrid",None,"Annual KPI","Value"]
    for c,v in enumerate(headers,1):
        cell=ws.cell(2,c,v); cell.fill=HEADER_FILL; cell.font=Font(color="FFFFFF",bold=True); cell.border=BORDER
    source_cols=["AI","AO","AU","AY","AZ","BC","BD","BH"]
    for day in range(365):
        out_r=3+day; start=2+day*24; end=start+23
        ws.cell(out_r,1,f"={HOURLY_SHEET}!B{start}")
        for j,col in enumerate(source_cols,2): ws.cell(out_r,j,f"=SUM({HOURLY_SHEET}!{col}{start}:{col}{end})")
        ws.cell(out_r,1).number_format="yyyy-mm-dd"
    kpis=[
        ("Annual Umax EQ8/29", "=SUM(B3:B367)"), ("Annual Uprosumer EQ14", "=SUM(C3:C367)"),
        ("Annual SWtotal EQ15", "=SUM(D3:D367)"), ("Annual CS EQ23", "=SUM(E3:E367)"),
        ("Annual PS EQ24", "=SUM(F3:F367)"), ("Annual SWP2P EQ25", "=SUM(G3:G367)"),
        ("Annual CSpros EQ28", "=SUM(H3:H367)"), ("Annual SWgrid EQ32", "=SUM(I3:I367)"),
        ("Annual charge (kWh)", f"=SUM({HOURLY_SHEET}!T2:T8761)"),
        ("Annual discharge (kWh)", f"=SUM({HOURLY_SHEET}!U2:U8761)"),
        ("Annual grid purchase (kWh)", f"=SUM({HOURLY_SHEET}!AG2:AG8761)"),
        ("Annual sale (kWh)", f"=SUM({HOURLY_SHEET}!AF2:AF8761)"),
        ("Minimum SOC (kWh)", f"=MIN({HOURLY_SHEET}!W2:W8761)"),
        ("Maximum SOC (kWh)", f"=MAX({HOURLY_SHEET}!W2:W8761)"),
        ("Equivalent full cycles", f"=SUM({HOURLY_SHEET}!U2:U8761)/({battery:g}*(Parameters!$C$13-Parameters!$C$12))"),
        ("SOC range utilization", f"=(MAX({HOURLY_SHEET}!W2:W8761)-MIN({HOURLY_SHEET}!W2:W8761))/({battery:g}*(Parameters!$C$13-Parameters!$C$12))"),
    ]
    for i,(label,formula) in enumerate(kpis,3): ws.cell(i,11,label); ws.cell(i,12,formula)
    for col,width in {"A":14,"B":20,"C":22,"D":20,"E":18,"F":18,"G":20,"H":20,"I":20,"J":3,"K":28,"L":20}.items(): ws.column_dimensions[col].width=width
    ws.freeze_panes="A3"


def create_workbook(input_data: Path, core: Path, battery: float, pv_area: float, eta: float,
                    strategy: str = "A", args: argparse.Namespace | None = None) -> Workbook:
    source=load_workbook(input_data,data_only=False,read_only=False)
    if MONTHLY_SHEET not in source.sheetnames or HOURLY_SHEET not in source.sheetnames:
        raise KeyError(f"Input Data must contain sheets {MONTHLY_SHEET!r} and {HOURLY_SHEET!r}")
    if source[HOURLY_SHEET].max_row < LAST_DATA_ROW or source[HOURLY_SHEET].max_column < 10:
        raise ValueError("Input hourly sheet must contain at least 8760 data rows and columns A:J")
    wb=Workbook(); wb.remove(wb.active)
    monthly=wb.create_sheet(MONTHLY_SHEET); hourly=wb.create_sheet(HOURLY_SHEET); params=wb.create_sheet(PARAMETERS_SHEET); guide=wb.create_sheet(GUIDE_SHEET); summary=wb.create_sheet(SUMMARY_SHEET)
    copy_input_sheet_values(source[MONTHLY_SHEET],monthly)
    copy_input_sheet_values(source[HOURLY_SHEET],hourly,max_col=10)
    monthly.freeze_panes="A2"; monthly.auto_filter.ref=f"A1:H{monthly.max_row}"
    for c in range(1,monthly.max_column+1):
        monthly.column_dimensions[get_column_letter(c)].width=22
        monthly.cell(1,c).fill=HEADER_FILL; monthly.cell(1,c).font=Font(color="FFFFFF",bold=True)
    create_parameters(params,battery,eta,input_data.name,core.name,strategy,args)
    create_guide(guide,battery)
    guide["A1"] = f"EQ1-EQ32 Guide — Battery {battery:g} kWh — Strategy {strategy}: {STRATEGY_NAMES[strategy]}"
    if strategy in ("A","B"):
        write_hourly_formulas_strategy_ab(hourly,pv_area,strategy)
    else:
        write_hourly_strategy_c(hourly,pv_area,battery,eta,args)
    create_summary(summary,battery,strategy)
    wb.calculation.fullCalcOnLoad=True; wb.calculation.forceFullCalc=True; wb.calculation.calcMode="auto"
    wb.active=1
    return wb


def verify_workbook(path: Path, battery: float, strategy: str) -> None:
    wb=load_workbook(path,data_only=False,read_only=True)
    for s in [MONTHLY_SHEET,HOURLY_SHEET,PARAMETERS_SHEET,GUIDE_SHEET,SUMMARY_SHEET]:
        if s not in wb.sheetnames: raise AssertionError(f"Missing sheet: {s}")
    if float(wb[PARAMETERS_SHEET]["C6"].value) != float(battery): raise AssertionError("Battery mismatch")
    if wb[PARAMETERS_SHEET]["C29"].value != STRATEGY_NAMES[strategy]: raise AssertionError("Strategy mismatch")
    if wb[HOURLY_SHEET]["BE2"].value is not None: raise AssertionError("BE must remain blank")


def main() -> int:
    args=parse_args()
    try:
        input_data=existing_or_fallback(args.input_data,["Input Data*.xlsx"])
        core=existing_or_fallback(args.core_assumptions,["Input Core assumptions*.xlsx"])
        pv_area,_,eta=read_core_assumptions(core)
        args.output_dir.mkdir(parents=True,exist_ok=True)
        sizes=[float(x) for x in args.battery_sizes]
        strategies=list(dict.fromkeys(args.strategies))
        outputs=[]
        for strategy in strategies:
            if strategy == "C" and linprog is None:
                raise RuntimeError("Strategy C selected but scipy/numpy are not installed. Run: pip install scipy numpy")
            for battery in sizes:
                label=f"{battery:g}".replace(".","p")
                out=args.output_dir/f"Output_EQ1_EQ32_Strategy_{strategy}_Battery_{label}kWh.xlsx"
                wb=create_workbook(input_data,core,battery,pv_area,eta,strategy,args)
                wb.save(out)
                verify_workbook(out,battery,strategy)
                outputs.append(out)
                print(f"Created: {out}")
        with zipfile.ZipFile(args.zip_output,"w",compression=zipfile.ZIP_DEFLATED) as zf:
            for out in outputs: zf.write(out,arcname=out.name)
        print(f"Created ZIP: {args.zip_output}")
        return 0
    except Exception as exc:
        print(f"ERROR: {exc}",file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
